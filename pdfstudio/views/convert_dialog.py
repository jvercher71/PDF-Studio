"""
Zeus PDF — Format Conversion Dialog

Supported output formats:
  PNG / JPEG   — page images via PyMuPDF
  SVG          — scalable vector via PyMuPDF
  HTML         — text+layout via PyMuPDF
  TXT          — plain text extraction
  DOCX         — via pdf2docx (installed on demand)
  XLSX         — table extraction via PyMuPDF, written with openpyxl
"""
import logging
import io
from pathlib import Path
from typing import Optional

import fitz
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QCheckBox, QSpinBox, QProgressBar, QFileDialog,
    QDialogButtonBox, QFrame, QGroupBox, QFormLayout,
    QMessageBox, QScrollArea, QWidget, QSizePolicy,
)

from pdfstudio.models.document_model import DocumentModel

log = logging.getLogger(__name__)

FORMATS = [
    ("PNG Image (*.png)",       "png"),
    ("JPEG Image (*.jpg)",      "jpg"),
    ("SVG Vector (*.svg)",      "svg"),
    ("HTML Document (*.html)",  "html"),
    ("Plain Text (*.txt)",      "txt"),
    ("Word Document (*.docx)",  "docx"),
    ("Excel Workbook (*.xlsx)", "xlsx"),
]


# ── Background conversion worker ────────────────────────────────────────
class ConversionWorker(QThread):
    progress   = Signal(int, int)    # current, total
    finished   = Signal(str)         # output path
    error      = Signal(str)         # error message

    def __init__(self, doc: fitz.Document, fmt: str,
                 output_path: str, dpi: int,
                 page_range: list[int], parent=None):
        super().__init__(parent)
        self._doc = doc
        self._fmt = fmt
        self._output = output_path
        self._dpi = dpi
        self._pages = page_range

    def run(self):
        try:
            match self._fmt:
                case "png" | "jpg":
                    self._export_images()
                case "svg":
                    self._export_svg()
                case "html":
                    self._export_html()
                case "txt":
                    self._export_txt()
                case "docx":
                    self._export_docx()
                case "xlsx":
                    self._export_xlsx()
                case _:
                    self.error.emit(f"Unknown format: {self._fmt}")
        except Exception as e:
            log.error("Conversion error: %s", e, exc_info=True)
            self.error.emit(str(e))

    # ── Image ─────────────────────────────────────────────────────────
    def _export_images(self):
        total = len(self._pages)
        zoom = self._dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        out = Path(self._output)
        is_single = total == 1

        for i, page_idx in enumerate(self._pages):
            page = self._doc[page_idx]
            pix = page.get_pixmap(matrix=mat, alpha=False)
            if is_single:
                dest = out
            else:
                dest = out.parent / f"{out.stem}_p{page_idx + 1:03d}{out.suffix}"

            if self._fmt == "jpg":
                pix.save(str(dest), jpg_quality=92)
            else:
                pix.save(str(dest))
            self.progress.emit(i + 1, total)

        self.finished.emit(str(out.parent if not is_single else out))

    # ── SVG ───────────────────────────────────────────────────────────
    def _export_svg(self):
        total = len(self._pages)
        out = Path(self._output)
        is_single = total == 1

        for i, page_idx in enumerate(self._pages):
            page = self._doc[page_idx]
            svg = page.get_svg_image(matrix=fitz.Identity)
            dest = out if is_single else \
                out.parent / f"{out.stem}_p{page_idx + 1:03d}.svg"
            dest.write_text(svg, encoding="utf-8")
            self.progress.emit(i + 1, total)

        self.finished.emit(str(out.parent if not is_single else out))

    # ── HTML ──────────────────────────────────────────────────────────
    def _export_html(self):
        total = len(self._pages)
        out = Path(self._output)
        chunks = [
            "<!DOCTYPE html><html><head>"
            "<meta charset='utf-8'>"
            "<title>Zeus PDF Export</title>"
            "<style>body{font-family:sans-serif;max-width:900px;margin:auto;padding:2em}"
            ".page{border-bottom:2px solid #eee;margin-bottom:2em;padding-bottom:2em}"
            ".page-num{color:#aaa;font-size:12px;margin-bottom:1em}</style>"
            "</head><body>"
        ]

        for i, page_idx in enumerate(self._pages):
            page = self._doc[page_idx]
            html = page.get_text("html")
            chunks.append(
                f"<div class='page'>"
                f"<div class='page-num'>Page {page_idx + 1}</div>"
                f"{html}</div>"
            )
            self.progress.emit(i + 1, total)

        chunks.append("</body></html>")
        out.write_text("".join(chunks), encoding="utf-8")
        self.finished.emit(str(out))

    # ── TXT ───────────────────────────────────────────────────────────
    def _export_txt(self):
        total = len(self._pages)
        out = Path(self._output)
        lines = []

        for i, page_idx in enumerate(self._pages):
            page = self._doc[page_idx]
            text = page.get_text("text")
            lines.append(f"{'='*60}\nPage {page_idx + 1}\n{'='*60}\n{text}\n")
            self.progress.emit(i + 1, total)

        out.write_text("\n".join(lines), encoding="utf-8")
        self.finished.emit(str(out))

    # ── DOCX ──────────────────────────────────────────────────────────
    def _export_docx(self):
        try:
            from pdf2docx import Converter
        except ImportError:
            self.error.emit(
                "pdf2docx is not installed.\n\n"
                "Install it with:\n  pip install pdf2docx\n\n"
                "Then try again."
            )
            return

        out = Path(self._output)
        total = len(self._pages)

        # pdf2docx works on a file path, not a fitz doc
        # We need to save a temp PDF of just the selected pages
        import tempfile, shutil
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Save selected pages to temp file
            new_doc = fitz.open()
            for page_idx in self._pages:
                new_doc.insert_pdf(self._doc, from_page=page_idx, to_page=page_idx)
            new_doc.save(tmp_path)
            new_doc.close()

            cv = Converter(tmp_path)
            cv.convert(str(out))
            cv.close()
            self.progress.emit(total, total)
            self.finished.emit(str(out))
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    # ── XLSX ──────────────────────────────────────────────────────────
    def _export_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.error.emit(
                "openpyxl is not installed.\n\n"
                "Install it with:\n  pip install openpyxl\n\nThen try again."
            )
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet
        total = len(self._pages)

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="0F1A2D")
        cell_align   = Alignment(wrap_text=True, vertical="top")
        thin_border  = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
        )

        for i, page_idx in enumerate(self._pages):
            page = self._doc[page_idx]
            ws = wb.create_sheet(title=f"Page {page_idx + 1}")

            # Extract tables via PyMuPDF structured text
            tables = page.find_tables()
            if tables and tables.tables:
                for t_idx, table in enumerate(tables.tables):
                    df = table.to_pandas() if hasattr(table, "to_pandas") else None
                    rows = table.extract()
                    if not rows:
                        continue
                    start_row = ws.max_row + (2 if ws.max_row > 1 else 1)
                    ws.cell(start_row, 1, f"Table {t_idx + 1}").font = Font(bold=True)
                    start_row += 1

                    for r_idx, row in enumerate(rows):
                        for c_idx, cell_val in enumerate(row):
                            c = ws.cell(start_row + r_idx, c_idx + 1, cell_val or "")
                            c.alignment = cell_align
                            c.border = thin_border
                            if r_idx == 0:
                                c.font = header_font
                                c.fill = header_fill
                    ws.append([])   # blank row between tables
            else:
                # No tables — dump text blocks line by line
                blocks = page.get_text("blocks")
                for block in blocks:
                    text = block[4].strip()
                    if text:
                        for line in text.split("\n"):
                            if line.strip():
                                ws.append([line.strip()])

            # Auto-fit column widths (rough estimate)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            self.progress.emit(i + 1, total)

        out = Path(self._output)
        wb.save(str(out))
        self.finished.emit(str(out))


# ── Dialog ──────────────────────────────────────────────────────────────
class ConvertDialog(QDialog):

    def __init__(self, model: DocumentModel, parent=None):
        super().__init__(parent)
        self._model = model
        self._worker: Optional[ConversionWorker] = None

        self.setWindowTitle("⚡ Zeus PDF — Convert / Export")
        self.setMinimumWidth(500)
        self.setModal(True)

        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        # ── Format selector ──────────────────────────────────────────
        fmt_group = QGroupBox("Output Format")
        fmt_form = QFormLayout(fmt_group)

        self._fmt_combo = QComboBox()
        for label, _ in FORMATS:
            self._fmt_combo.addItem(label)
        self._fmt_combo.currentIndexChanged.connect(self._on_format_changed)
        fmt_form.addRow("Format:", self._fmt_combo)
        layout.addWidget(fmt_group)

        # ── Page range ───────────────────────────────────────────────
        range_group = QGroupBox("Pages")
        range_form = QFormLayout(range_group)

        self._range_combo = QComboBox()
        self._range_combo.addItems(["All Pages", "Current Page", "Custom Range"])
        self._range_combo.currentIndexChanged.connect(self._on_range_changed)
        range_form.addRow("Export:", self._range_combo)

        self._range_row = QHBoxLayout()
        self._from_spin = QSpinBox()
        self._from_spin.setMinimum(1)
        self._from_spin.setMaximum(self._model.page_count)
        self._to_spin = QSpinBox()
        self._to_spin.setMinimum(1)
        self._to_spin.setMaximum(self._model.page_count)
        self._to_spin.setValue(self._model.page_count)
        self._range_row.addWidget(QLabel("From page:"))
        self._range_row.addWidget(self._from_spin)
        self._range_row.addWidget(QLabel("  To:"))
        self._range_row.addWidget(self._to_spin)
        self._range_row.addStretch()
        range_form.addRow("", self._range_row)
        self._toggle_range_row(False)
        layout.addWidget(range_group)

        # ── Image quality ────────────────────────────────────────────
        self._img_group = QGroupBox("Image Quality")
        img_form = QFormLayout(self._img_group)
        self._dpi_spin = QSpinBox()
        self._dpi_spin.setRange(72, 600)
        self._dpi_spin.setValue(150)
        self._dpi_spin.setSuffix(" DPI")
        img_form.addRow("Resolution:", self._dpi_spin)
        layout.addWidget(self._img_group)

        # ── Info label ───────────────────────────────────────────────
        self._info_label = QLabel()
        self._info_label.setWordWrap(True)
        self._info_label.setStyleSheet("color: #9D9D9D; font-size: 12px;")
        layout.addWidget(self._info_label)
        self._update_info()

        # ── Progress ─────────────────────────────────────────────────
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        self._progress.setTextVisible(True)
        layout.addWidget(self._progress)

        # ── Buttons ──────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self._convert_btn = QPushButton("⚡  Convert & Save…")
        self._convert_btn.setDefault(True)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        self._convert_btn.clicked.connect(self._on_convert)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(self._convert_btn)
        layout.addLayout(btn_row)

        self._on_format_changed(0)

    # ── Helpers ──────────────────────────────────────────────────────

    def _on_format_changed(self, _):
        fmt = self._current_fmt()
        self._img_group.setVisible(fmt in ("png", "jpg"))
        self._update_info()

    def _on_range_changed(self, idx):
        self._toggle_range_row(idx == 2)

    def _toggle_range_row(self, visible: bool):
        for i in range(self._range_row.count()):
            item = self._range_row.itemAt(i)
            if item and item.widget():
                item.widget().setVisible(visible)

    def _update_info(self):
        infos = {
            "png":  "Exports each page as a PNG image. Multiple pages → numbered files.",
            "jpg":  "Exports each page as a JPEG image (92% quality).",
            "svg":  "Exports pages as scalable vector graphics.",
            "html": "Exports text and basic layout as an HTML file.",
            "txt":  "Extracts all text content as plain text.",
            "docx": "Converts to Word format. Requires: pip install pdf2docx",
            "xlsx": "Extracts tables/text into Excel. Requires: pip install openpyxl",
        }
        self._info_label.setText(infos.get(self._current_fmt(), ""))

    def _current_fmt(self) -> str:
        return FORMATS[self._fmt_combo.currentIndex()][1]

    def _page_range(self) -> list[int]:
        mode = self._range_combo.currentIndex()
        total = self._model.page_count
        if mode == 0:
            return list(range(total))
        elif mode == 1:
            # current page — get from parent if possible
            parent = self.parent()
            cur = getattr(getattr(parent, "_canvas", None), "_current_page", 0)
            return [cur]
        else:
            f = self._from_spin.value() - 1
            t = self._to_spin.value() - 1
            return list(range(max(0, f), min(total, t + 1)))

    def _on_convert(self):
        fmt = self._current_fmt()
        ext_map = dict(FORMATS)
        # Build save dialog filter from selected format
        save_filter = FORMATS[self._fmt_combo.currentIndex()][0]
        output, _ = QFileDialog.getSaveFileName(
            self, "Save As", "", save_filter
        )
        if not output:
            return

        # Ensure extension
        p = Path(output)
        if p.suffix.lower().lstrip(".") != fmt:
            output = str(p.with_suffix(f".{fmt}"))

        pages = self._page_range()
        doc = self._model._pdf.raw()

        self._convert_btn.setEnabled(False)
        self._progress.setVisible(True)
        self._progress.setMaximum(len(pages))
        self._progress.setValue(0)

        self._worker = ConversionWorker(
            doc, fmt, output, self._dpi_spin.value(), pages, parent=self
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_done)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_progress(self, current: int, total: int):
        self._progress.setValue(current)
        self._progress.setFormat(f"Converting page {current} of {total}…")

    def _on_done(self, output_path: str):
        self._progress.setVisible(False)
        self._convert_btn.setEnabled(True)
        QMessageBox.information(
            self, "Conversion Complete",
            f"✅  Saved to:\n{output_path}"
        )
        self.accept()

    def _on_error(self, msg: str):
        self._progress.setVisible(False)
        self._convert_btn.setEnabled(True)
        QMessageBox.critical(self, "Conversion Failed", msg)
